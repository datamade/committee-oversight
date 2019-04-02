from sys import stdout, argv
from openpyxl import load_workbook, writer

name_index = argv[1]
file_path = argv[2]
new_path = argv[3]

# load excel file and name first sheet
wb = load_workbook(file_path)
ws = wb['Hearings']

# insert column at 0
ws.insert_cols(0)

# name column
ws['A1'] = "source"

# fill column with hyperlink value from name column
for row in ws.rows:
    link = row[int(name_index)].hyperlink
    if link:
        row[0].value = link.target
# save workbook
wb.save(new_path)
